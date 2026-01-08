"""Reporting system for DirPurge operations"""

import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List
from .models import PurgeResult, Config

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


class Reporter:
    """Generates detailed reports of purge operations"""
    
    def __init__(self, reports_directory: str = "./reports"):
        """Initialize reporter with reports directory"""
        self.reports_directory = Path(reports_directory)
    
    def generate_report(self, purge_result: PurgeResult, config: Config, file_groups: Dict[str, List]) -> str:
        """Generate detailed report content"""
        report_lines = []
        
        # Header
        report_lines.append("=" * 60)
        report_lines.append("DirPurge Operation Report")
        report_lines.append("=" * 60)
        report_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report_lines.append(f"Target Directory: {config.target_directory}")
        report_lines.append(f"Dry Run Mode: {'Yes' if config.dry_run else 'No'}")
        report_lines.append("")
        
        # Configuration
        report_lines.append("Configuration:")
        report_lines.append(f"  Minimum files to keep: {config.min_files_to_keep}")
        report_lines.append(f"  Maximum age (days): {config.max_age_days}")
        report_lines.append(f"  Excluded extensions: {', '.join(config.excluded_extensions)}")
        report_lines.append("")
        
        # Summary statistics
        report_lines.append("Summary:")
        report_lines.append(f"  Total files scanned: {purge_result.total_files_scanned}")
        report_lines.append(f"  File sets found: {purge_result.file_sets_found}")
        report_lines.append(f"  Files to delete: {len(purge_result.files_to_delete)}")
        report_lines.append(f"  Files preserved: {len(purge_result.files_preserved)}")
        report_lines.append(f"  Actual deletions: {purge_result.actual_deletions}")
        report_lines.append(f"  Execution time: {purge_result.execution_time:.2f} seconds")
        report_lines.append("")
        
        # File set details
        if file_groups:
            report_lines.append("File Set Details:")
            for prefix, files in file_groups.items():
                files_to_delete_count = sum(1 for f in purge_result.files_to_delete if f.prefix == prefix)
                files_preserved_count = sum(1 for f in purge_result.files_preserved if f.prefix == prefix)
                
                report_lines.append(f"  {prefix}:")
                report_lines.append(f"    Total files: {len(files)}")
                report_lines.append(f"    To delete: {files_to_delete_count}")
                report_lines.append(f"    Preserved: {files_preserved_count}")
            report_lines.append("")
        
        # Errors
        if purge_result.errors:
            report_lines.append("Errors:")
            for error in purge_result.errors:
                report_lines.append(f"  - {error}")
            report_lines.append("")
        
        # Files to delete (if not too many)
        if purge_result.files_to_delete and len(purge_result.files_to_delete) <= 50:
            report_lines.append("Files to Delete:")
            for file_info in purge_result.files_to_delete:
                report_lines.append(f"  - {file_info.name} ({file_info.modified_time.strftime('%Y-%m-%d %H:%M:%S')})")
            report_lines.append("")
        elif len(purge_result.files_to_delete) > 50:
            report_lines.append(f"Files to Delete: {len(purge_result.files_to_delete)} files (list truncated)")
            report_lines.append("")
        
        return "\n".join(report_lines)
    
    def save_report(self, report_content: str, timestamp: datetime = None) -> Path:
        """Save report to file with YYYYMMDD-HH-MM format"""
        if timestamp is None:
            timestamp = datetime.now()
        
        # Create reports directory if it doesn't exist
        self.reports_directory.mkdir(parents=True, exist_ok=True)
        
        # Generate filename with YYYYMMDD-HH-MM format
        filename = self.format_filename(timestamp)
        report_path = self.reports_directory / filename
        
        # Write report to file
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write(report_content)
        
        return report_path
    
    def format_filename(self, timestamp: datetime) -> str:
        """Format filename using YYYYMMDD-HH-MM format"""
        return f"dirpurge_{timestamp.strftime('%Y%m%d-%H-%M')}.txt"
    
    def generate_xlsx_report(self, file_groups: Dict[str, List], timestamp: datetime = None) -> Path:
        """Generate Excel file with file set summary"""
        if not XLSX_AVAILABLE:
            raise ImportError("openpyxl library is required for XLSX export")
        
        if timestamp is None:
            timestamp = datetime.now()
        
        # Create reports directory if it doesn't exist
        self.reports_directory.mkdir(parents=True, exist_ok=True)
        
        # Generate filename with YYYYMMDD-HH-MM format
        filename = f"dirpurge_{timestamp.strftime('%Y%m%d-%H-%M')}.xlsx"
        xlsx_path = self.reports_directory / filename
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "File Sets Summary"
        
        # Add headers
        headers = ["Set Name", "Newest File Date", "Total Files"]
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data rows
        for set_name, files in file_groups.items():
            if files:
                # Find the newest file in the set
                newest_file = max(files, key=lambda f: f.modified_time)
                newest_date = newest_file.modified_time.strftime('%Y-%m-%d %H:%M:%S')
                total_files = len(files)
                
                ws.append([set_name, newest_date, total_files])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        wb.save(xlsx_path)
        
        return xlsx_path